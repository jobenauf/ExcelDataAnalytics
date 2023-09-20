import os
import openpyxl
import concurrent.futures

# Get the list of Excel files ending with "output.xlsx" in the directory
excel_files = [filename for filename in os.listdir('.') if filename.endswith('.xlsx')]

# Create a dictionary to store data for each participant ID
participant_data = {}

# Create a list to store unique data types in the order they appear
unique_data_types = []

# Loop through all Excel files
for excel_file in excel_files:
    data_workbook = openpyxl.load_workbook(excel_file, data_only=True)
    data_sheet = data_workbook.active

    try:
        # Get the first row (headers) from the sheet
        headers = next(data_sheet.iter_rows(values_only=True))
    except StopIteration:
        # Skip empty sheets
        continue

    # If it's the first sheet, collect unique data types and their order
    if not unique_data_types:
        unique_data_types.extend(headers[1:])  # Exclude the ID column
    else:
        # Update unique_data_types with any new data types
        for header in headers[1:]:
            if header not in unique_data_types:
                unique_data_types.append(header)

    # Process the rest of the rows
    for row in data_sheet.iter_rows(values_only=True):
        participant_id = row[0]  # Assuming participant ID is in the first column

        if participant_id not in participant_data:
            participant_data[participant_id] = {}

        # Populate data for the corresponding data types
        for header, value in zip(headers[1:], row[1:]):
            participant_data[participant_id][header] = value

# Create a new workbook for the combined data
combined_workbook = openpyxl.Workbook()
combined_sheet = combined_workbook.active

# Create the first row with unique data types in their original order
combined_sheet.append(['ID'] + unique_data_types)

# Populate the rest of the rows with participant data
for participant_id, data in participant_data.items():
    row_data = [participant_id] + [data.get(data_type, None) for data_type in unique_data_types]
    combined_sheet.append(row_data)

# Save the combined workbook
combined_workbook.save('combined.xlsx')
